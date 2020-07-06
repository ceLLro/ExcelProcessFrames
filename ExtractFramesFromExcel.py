# usage : python3 scriptName.py "csv file"(without quotes {""}). results folder == where python script is. this
# script will take all the data available in column 10 (fovio.results_generation_utc_time.time_s) if
# fovio.results_generation_utc_time.time_s have mapped another letter (not J which is 10, in excel file),
# modify min_col & max_col from line 24 accordingly.

import sys
import os
import shutil
from openpyxl import load_workbook
from pathlib import Path
from shutil import copyfile
import time, datetime
from datetime import datetime

def get_frame_seconds(path):
    wb = load_workbook(filename=path)
    ws = wb.active
    file = open(filename + '.txt', "w")

    for col in ws.iter_cols(min_row=2, min_col=10, max_col=10, max_row=ws.max_row):
        for cell in col:
            file.write(str(cell.value) + '\n')
    file.write('99999')

    file.close()
    file = open(source + 'seconds.txt', "w")
    with open(filename + '.txt', 'r') as f:
        lines = f.readlines()
    numbers = [int(e.strip()) for e in lines]
    numbers.sort()
    for l in numbers:
        file.write('.' + str(l) + '\n')
    file.close()
    os.remove(filename + '.txt')

def get_framerate(path):
    fp = open(path, 'r')
    list_of_seconds = fp.readlines()
    finalList = [int(sub.split('.')[1]) for sub in list_of_seconds]
    counter = 0
    listArray = []
    averageArray = []
    average = 0
    for i in range(0, len(finalList)):
        if finalList[i] == 99999:
            print("\n")
        else:
            if finalList[i] == finalList[i + 1]:
                counter = counter + 1
            else:
                listArray.append(str(finalList[i]) + " - " + str(counter + 1))
                averageArray.append(counter + 1)
                counter = 0

    for i in range(0, len(averageArray)):
        average = average + averageArray[i]
        totalAverage = average / len(averageArray)

    f = open(source + filename + "_output.txt", "w")
    f.write("Total frames : " + str(len(finalList) - 1) + "\n\n")
    f.write("Average frame for each second : " + str(listArray) + "\n\n")
    f.write("Average frame : " + str(averageArray) + "\n\n")
    f.write("Average FPS: " + str(totalAverage) + "\n\n")
    f.close()
    fp.close()

    # create report folder
    now = datetime.now()
    dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")

    repPath = destination + dt_string + '_' + filename
    os.makedirs(repPath)

    shutil.copy(pathToFile, repPath)
    shutil.move(source + 'seconds.txt', repPath)
    shutil.move(source + filename + '_output.txt', repPath)


# this will return the number of "\" in a filepath
def returnDashes(txt):
    delim = 0
    for i in txt:
        if i == '\\':
            delim = delim + 1
    return delim


# if no argument is provided, return exit;
if len(sys.argv) != 2:
    print("Provide 1 argument")
    print("Usage:")
    print("python python_file \"excel_File\"")
    print("EG: python ExtractFramesFromExcel.py \"C:\\Users\\user.name\\Desktop\\example.xsls\"")
    sys.exit()

# path to excel file
pathToFile = str(sys.argv[1])

# file name
filename = pathToFile.split("\\", returnDashes(pathToFile))
filename = filename[returnDashes(pathToFile)]

# path where the file was provided
source = pathToFile.replace(filename, '')

# destination reports
destination = source + 'Results\\'

get_frame_seconds(pathToFile)
get_framerate(source + 'seconds.txt');
